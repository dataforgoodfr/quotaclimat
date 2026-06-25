"""Parse a Pulsar TRAC "Themes" export workbook.

Two sheets:
  - "Themes": one row per theme — [topics-string, Title, Summary, Volume, Sentiment].
    The first column is the theme's constituent-topics string, which also keys the
    Topics Breakdown sheet.
  - "Topics Breakdown": one row per topic — [Topic, Volume, theme-topics-string].
    We join each topic to its theme via that shared topics-string.

The preamble holds the search name and the pull's date window.
"""

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

_PARIS = ZoneInfo("Europe/Paris")
_DATE_RE = re.compile(r"(\d{2}-\d{2}-\d{4})\s*\((\d{2}:\d{2})\)")


@dataclass
class Topic:
    label: str
    volume: int = 0


@dataclass
class Theme:
    title: str
    summary: str = ""
    volume: int = 0
    sentiment: str = ""
    topics_key: str = ""  # the constituent-topics string (join key)
    topics: list[Topic] = field(default_factory=list)


@dataclass
class ThemesExport:
    search_name: str
    date_from: datetime | None
    date_to: datetime | None
    themes: list[Theme]


def _to_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value).replace(",", "").strip()))
    except ValueError:
        return 0


def _parse_dt(value) -> datetime | None:
    m = _DATE_RE.search(str(value or ""))
    if not m:
        return None
    return datetime.strptime(f"{m.group(1)} {m.group(2)}", "%d-%m-%Y %H:%M").replace(tzinfo=_PARIS)


def _rows(wb, sheet) -> list:
    if sheet not in wb.sheetnames:
        return []
    return [r for r in wb[sheet].iter_rows(values_only=True)]


def _header_idx(rows, first_two) -> int | None:
    return next(
        (
            i
            for i, r in enumerate(rows)
            if r and [str(c).strip() if c is not None else "" for c in r[:2]] == list(first_two)
        ),
        None,
    )


def parse_themes_xlsx(path) -> ThemesExport:
    wb = openpyxl.load_workbook(Path(path), read_only=True, data_only=True)
    try:
        theme_rows = _rows(wb, "Themes")
        topic_rows = _rows(wb, "Topics Breakdown")
    finally:
        wb.close()

    search_name, date_from, date_to = "", None, None
    for r in theme_rows[:8]:
        if not r:
            continue
        key = str(r[0] or "").strip().lower()
        val = r[1] if len(r) > 1 else None
        if key == "search name":
            search_name = str(val or "").strip()
        elif key == "date from":
            date_from = _parse_dt(val)
        elif key == "date to":
            date_to = _parse_dt(val)

    # Topics Breakdown → {theme_key: [Topic, ...]}
    topics_by_key: dict[str, list[Topic]] = {}
    t_hdr = _header_idx(topic_rows, ("Topics", "Volume of Posts"))
    if t_hdr is not None:
        for r in topic_rows[t_hdr + 1 :]:
            if not r or not r[0]:
                continue
            label, volume, theme_key = (list(r) + [None] * 3)[:3]
            topics_by_key.setdefault(str(theme_key or "").strip(), []).append(
                Topic(label=str(label).strip(), volume=_to_int(volume))
            )

    themes: list[Theme] = []
    h = _header_idx(theme_rows, ("Themes", "Title"))
    if h is None:
        raise ValueError(f"'Themes' header row not found in {path}")
    for r in theme_rows[h + 1 :]:
        if not r or all(c is None for c in r):
            continue
        topics_key, title, summary, volume, sentiment = (list(r) + [None] * 5)[:5]
        if not title:
            continue
        key = str(topics_key or "").strip()
        themes.append(
            Theme(
                title=str(title).strip(),
                summary=str(summary or "").strip(),
                volume=_to_int(volume),
                sentiment=str(sentiment or "").strip(),
                topics_key=key,
                topics=topics_by_key.get(key, []),
            )
        )

    return ThemesExport(search_name=search_name, date_from=date_from, date_to=date_to, themes=themes)
